
import smtplib
import time
import openpyxl
from email.message import EmailMessage
#NomeVariavelPlanilha = openpyxl.load_workbook('PlanilhaQueDeseja.xlsx',data_only='false')
wb = openpyxl.load_workbook('controleRenovacao.xlsx',data_only='false')
#Variavel da aba = variavelDaPlan['AbaDaPlanilhaQueDeseja']
sheet = wb['DiasParaVencer']
row_count = sheet.max_row
column_count = sheet.max_column
for i in range(1, row_count):
    for j in range(1, column_count):
        #Licenças
        #NomeDaVariavel = sheet.cell(row=numeroDaLinha, column=NumeroDaColuna).value Value é para pegar o valor da celula
        nomeColuna = str(sheet.cell(row=1, column=3).value)
        #str() - para fazer leitura do dado da variavel como string
        #int() - para fazer leitura do dado da variavel como inteiro
        licencaIndividualTeams = int(sheet.cell(row=2, column=3).value)
        licenca16Teams = int(sheet.cell(row=3, column=3).value)
        licenca18Teams = int(sheet.cell(row=4, column=3).value)
        licencaBackupExec = int(sheet.cell(row=5, column=3).value)
        licencaTeamViewer = int(sheet.cell(row=6, column=3).value)
        licencaAdobePRO = int(sheet.cell(row=9, column=3).value)
        #Certificados
        nomeColuna2 = sheet.cell(row=1, column=3).value
        CertificadoA1 = int(sheet.cell(row=9, column=3).value)
        CertificadoRemoteApp = int(sheet.cell(row=10, column=3).value)
        #Dominios
        nomeColuna3 = sheet.cell(row=1, column=3).value
        Domainusas = int(sheet.cell(row=11, column=3).value)
        Domainusos = int(sheet.cell(row=12, column=3).value)
        Domainused = int(sheet.cell(row=13, column=3).value)
        Domainusus = int(sheet.cell(row=14, column=3).value)

file = open('LogDiarioStatusLicenciamentos_Certificados_Dominios.txt', 'w')
file.write('\nLicenciamento/Assinaturas: ' + nomeColuna + '\n')
file.write('Faltam ' + str(licencaIndividualTeams) + ' dias para vencer a licença individual do Teams\n')
file.write('Faltam ' + str(licenca16Teams) + ' dias para vencer as 16 licenças do Teams \n')
file.write('Faltam ' + str(licenca18Teams) + ' dias para vencer as 18 licenças do Teams \n')
file.write('Faltam ' + str(licencaBackupExec) + ' dias para vencer a licença do BackupExec \n')
file.write('Faltam ' + str(licencaTeamViewer) + ' dias para vencer a licença do Team Viewer\n')
file.write('Faltam ' + str(licencaAdobePRO) + ' dias para vencer a licença do Oracle')
file.write('\n')
file.write('\nCertificados: ' + nomeColuna2 + '\n')
file.write('Faltam ' + str(CertificadoA1) + ' dias para vencer o certificado A1-Linck Maquinas\n')
file.write('Faltam ' + str(CertificadoRemoteApp) + ' dias para vencer o Certificado do Remote App TS4 e TS2\n')
file.write('\n')
file.write('Dominios: ' + nomeColuna3 + ' no RegistroBR\n')
file.write('Faltam ' + str(Domainusas) + ' dias para vencer o dominio usas.com.br\n')
file.write('Faltam ' + str(Domainusos) + ' dias para vencer o dominio usos.com.br\n')
file.write('Faltam ' + str(Domainused) + ' dias para vencer o dominio used.com.br\n')
file.write('Faltam ' + str(Domainusus) + ' dias para vencer o dominio usus.com.br\n')
file.close()

# Impressões de mensagem dos resultados
print('\nLicenciamento/Assinaturas: ' + nomeColuna)
print('Faltam ' + str(licencaIndividualTeams) + ' dias para vencer a licença individual do Teams')
print('Faltam ' + str(licenca16Teams) + ' dias para vencer as 16 licenças do Teams ')
print('Faltam ' + str(licenca18Teams) + ' dias para vencer as 18 licenças do Teams ')
print('Faltam ' + str(licencaBackupExec) + ' dias para vencer a licença do BackupExec ')
print('Faltam ' + str(licencaTeamViewer) + ' dias para vencer a licença do Team Viewer')
print('Faltam ' + str(licencaAdobePRO) + ' dias para vencer a licença do Adobe PRO - Diretor')
print()
print('Certificados: ' + nomeColuna2)
print('Faltam ' + str(CertificadoA1) + ' dias para vencer o certificado A1-Linck Maquinas')
print('Faltam ' + str(CertificadoRemoteApp) + ' dias para vencer o Certificado do Remote App TS4 e TS2')
print()
print('Dominios: ' + nomeColuna3 + ' no RegistroBR')
print('Faltam ' + str(Domainusas) + ' dias para vencer o dominio usas.com.br')
print('Faltam ' + str(Domainusos) + ' dias para vencer o dominio usos.com.br')
print('Faltam ' + str(Domainused) + ' dias para vencer o dominio used.com.br')
print('Faltam ' + str(Domainusus) + ' dias para vencer o dominio usus.com.br')
time.sleep(5)
# AutenticaçãoEmail
email_from = '***************@*************.com.br'
email_to = '********@*********.com.br'
smtp = 'mail.*************.com.br'
senha = '************'
# licenças
if (licencaIndividualTeams == 40 or licencaIndividualTeams == 30 or licencaIndividualTeams == 20 or
        licencaIndividualTeams == 15 or licencaIndividualTeams == 5 or licencaIndividualTeams == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Faltam ' + str(licencaIndividualTeams) + ' dias para vencer a licença individual do Teams')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (licenca16Teams == 40 or licenca16Teams == 30 or licenca16Teams == 20 or
        licenca16Teams == 15 or licenca16Teams == 5 or licenca16Teams == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(licenca16Teams) + ' dia(as) expiram as 16 licenças do Teams ')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (licenca18Teams == 40 or licenca18Teams == 30 or licenca18Teams == 20 or
        licenca18Teams == 15 or licenca18Teams == 5 or licenca18Teams == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(licenca18Teams) + ' dia(as) expiram as 18 licenças do Teams ')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (licencaBackupExec == 40 or licencaBackupExec == 30 or licencaBackupExec == 20 or
        licencaBackupExec == 15 or licencaBackupExec == 5 or licencaBackupExec == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(licencaBackupExec) + ' dia(as) expira a licença do BackupExec ')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (licencaTeamViewer == 40 or licencaTeamViewer == 30 or licencaTeamViewer == 20 or
        licencaTeamViewer == 15 or licencaTeamViewer == 5 or licencaTeamViewer == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Sua licença de Team Viewer expira em ' + str(licencaTeamViewer) + ' dia(as)')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (licencaAdobePRO == 40 or licencaAdobePRO == 30 or licencaAdobePRO == 20 or
        licencaAdobePRO == 15 or licencaAdobePRO == 5 or licencaAdobePRO == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(licencaAdobePRO) + ' dia(as) vence a licença de Adobe PRO - Diretor')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

# Certificados
if (CertificadoA1 == 40 or CertificadoA1 == 30 or CertificadoA1 == 20 or
        CertificadoA1 == 15 or CertificadoA1 == 5 or CertificadoA1 == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(CertificadoA1) + ' dias vence o certificado A1-Linck Maquinas')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (CertificadoRemoteApp == 40 or CertificadoRemoteApp == 30 or CertificadoRemoteApp == 20 or
        CertificadoRemoteApp == 15 or CertificadoRemoteApp == 5 or CertificadoRemoteApp == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(CertificadoRemoteApp) + ' dias vence o certificado do Remote App TS4 e TS2 ')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

# Dominios no Registro BR
if (Domainusas == 40 or Domainusas == 30 or Domainusas == 20 or
        Domainusas == 15 or Domainusas == 5 or Domainusas == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content(
        'Em ' + str(Domainusas) + ' dias vence o registro do dominio usas.com.br no RegistroBR')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (Domainusos == 40 or Domainusos == 30 or Domainusos == 20 or
        Domainusos == 15 or Domainusos == 5 or Domainusos == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content('Em ' + str(
        Domainusos) + ' dias vence registro do dominio usos.com.br no RegistroBR')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (Domainused == 40 or Domainused == 30 or Domainused == 20 or
        Domainused == 15 or Domainused == 5 or Domainused == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content(
        'Em ' + str(Domainused) + ' dias vence o registro do dominio used.com.br no RegistroBR')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)

if (Domainusus == 40 or Domainusus == 30 or Domainusus == 20 or
        Domainusus == 15 or Domainusus == 5 or Domainusus == 1):
    server = smtplib.SMTP(smtp, 587)
    server.starttls()
    server.login(email_from, senha)
    msg = EmailMessage()
    msg['Subject'] = 'ALERTA: Renovação Se Aproximando(Licença/Certificado/Dominio)'
    msg['From'] = email_from
    msg['To'] = email_to
    msg.set_content(
        'Em ' + str(Domainusus) + ' dias vence o registro do dominio USUS.COM.BR no RegistroBR')
    server.send_message(msg)
    server.quit()
    print('Email Enviado')
    time.sleep(1)